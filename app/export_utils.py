from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
from io import BytesIO

def export_to_pdf(data, title, filename=None):
    """
    Экспорт данных в PDF
    """
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    story = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#2C3E50'),
        spaceAfter=30
    )
    
    story.append(Paragraph(title, title_style))
    story.append(Spacer(1, 0.2*inch))
    
    if isinstance(data, list) and len(data) > 0:
        # Создаем таблицу
        table_data = []
        
        # Заголовки
        if isinstance(data[0], dict):
            headers = list(data[0].keys())
            table_data.append([str(h).replace('_', ' ').title() for h in headers])
            
            # Данные
            for row in data:
                table_data.append([str(row.get(h, '')) for h in headers])
        else:
            table_data = data
        
        table = Table(table_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2C3E50')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
        ]))
        
        story.append(table)
    
    doc.build(story)
    buffer.seek(0)
    return buffer

def export_to_excel(data, title, filename=None):
    """
    Экспорт данных в Excel
    """
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]  # Excel sheet name limit
    
    # Заголовок
    ws['A1'] = title
    ws['A1'].font = Font(bold=True, size=14, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    if isinstance(data, list) and len(data) > 0:
        # Заголовки
        if isinstance(data[0], dict):
            headers = list(data[0].keys())
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=3, column=col_idx)
                cell.value = str(header).replace('_', ' ').title()
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
                cell.alignment = Alignment(horizontal='center')
            
            # Данные
            for row_idx, row in enumerate(data, start=4):
                for col_idx, header in enumerate(headers, start=1):
                    ws.cell(row=row_idx, column=col_idx, value=str(row.get(header, '')))
        
        # Автоподбор ширины столбцов
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


